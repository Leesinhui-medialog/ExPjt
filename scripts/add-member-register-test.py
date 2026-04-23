# -*- coding: utf-8 -*-
"""
화면이동테스트결과서에 회원가입 테스트 케이스 추가 + '회원가입 캡처' 시트 생성.
이미지가 겹치지 않도록 실제 이미지 높이 기반으로 행 간격을 계산한다.
"""
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

wb = load_workbook("doc/화면이동테스트결과서.xlsx")

# ===== 1. 테스트 케이스 추가 =====
ws = wb["화면이동테스트결과서"]

data_font = Font(name="맑은 고딕", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

next_row = ws.max_row + 1

new_cases = [
    {
        "id": "SCT-F-MEMB-001", "type": "기능", "section": "화면이동",
        "major": "회원", "minor": "Step 01 약관동의",
        "reqId": "", "funcId": "MEMB-SCR-001", "progId": "MemberRegister.jsx", "screenId": "/member/register",
        "purpose": "회원가입 Step 01 약관동의 화면 정상 표시 및 동의 처리 검증",
        "procedure": "1. 로그인 화면에서 '회원가입' 클릭\n2. 이용약관, 개인정보 수집 동의 체크\n3. '다음' 버튼 클릭",
        "detail": "약관 내용이 표시되고, 두 항목 모두 체크 후 Step 02로 정상 이동",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-MEMB-002", "type": "기능", "section": "화면이동",
        "major": "회원", "minor": "Step 02 실명확인",
        "reqId": "", "funcId": "MEMB-SCR-002", "progId": "MemberRegister.jsx", "screenId": "/member/register",
        "purpose": "회원가입 Step 02 실명확인 화면 및 실명인증 처리 검증",
        "procedure": "1. 이름: '테스트회원' 입력\n2. 생년월일: '1995-06-15' 입력\n3. '실명인증' 버튼 클릭\n4. 인증 완료 확인 후 '다음' 클릭",
        "detail": "실명인증 완료 메시지 표시 후 Step 03으로 정상 이동",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-MEMB-003", "type": "기능", "section": "화면이동",
        "major": "회원", "minor": "Step 03 회원정보입력",
        "reqId": "", "funcId": "MEMB-SCR-003", "progId": "MemberRegister.jsx", "screenId": "/member/register",
        "purpose": "회원가입 Step 03 회원정보 입력 및 가입 완료 검증",
        "procedure": "1. 이메일: testai@medialog.co.kr 입력\n2. 전화번호: 010-9876-5432 입력\n3. 비밀번호/확인: test1234 입력\n4. '가입완료' 버튼 클릭",
        "detail": "회원가입 완료 팝업 표시 후 회원 목록 화면으로 자동 이동",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

for tc in new_cases:
    col_map = {
        2: tc["id"], 3: tc["type"], 4: tc["section"],
        5: tc["major"], 6: tc["minor"], 7: tc["reqId"],
        8: tc["funcId"], 9: tc["progId"], 10: tc["screenId"],
        11: tc["purpose"], 12: tc["procedure"], 13: tc["detail"],
        14: tc["date"], 15: tc["tester"], 16: tc["result"], 17: tc["reviewer"],
    }
    for col, value in col_map.items():
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)
    next_row += 1

# ===== 2. '회원가입 캡처' 시트 생성 =====
if "회원가입 캡처" in wb.sheetnames:
    del wb["회원가입 캡처"]

ws2 = wb.create_sheet("회원가입 캡처")

title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
c_align = Alignment(horizontal="center", vertical="center")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 100

# 이미지 목록: (파일명, 제목, 설명)
images = [
    ("09_회원가입_Step01_약관동의.png", "1. Step 01 약관동의 — 초기 화면", "이용약관 및 개인정보 수집 동의 화면"),
    ("10_회원가입_Step01_약관동의완료.png", "2. Step 01 약관동의 — 동의 완료", "이용약관, 개인정보 수집 모두 체크 완료 상태"),
    ("11_회원가입_Step02_실명확인.png", "3. Step 02 실명확인 — 초기 화면", "이름, 생년월일 입력 및 실명인증 버튼 표시"),
    ("12_회원가입_Step02_실명인증완료.png", "4. Step 02 실명확인 — 인증 완료", "실명인증 완료 메시지 표시 상태"),
    ("13_회원가입_Step03_회원정보입력.png", "5. Step 03 회원정보입력 — 초기 화면", "이메일, 전화번호, 비밀번호 입력 폼 표시"),
    ("14_회원가입_Step03_정보입력완료.png", "6. Step 03 회원정보입력 — 입력 완료", "모든 회원정보 입력 완료 상태"),
    ("15_회원가입_완료팝업.png", "7. 회원가입 완료", "가입완료 팝업 메시지 표시"),
]

# 이미지 너비 700px → 높이 비율 계산: 1280x900 → 700x492
IMG_WIDTH = 700
IMG_HEIGHT = 492
# 행 높이 기본 15px, 이미지 높이 492px → 약 33행
# 제목 2행 + 이미지 33행 + 여백 2행 = 37행 간격
ROW_GAP = 37

current_row = 2
for filename, title, desc in images:
    ws2.cell(row=current_row, column=2, value=title)
    ws2.cell(row=current_row, column=2).font = copy.copy(title_font)
    ws2.cell(row=current_row, column=2).alignment = copy.copy(c_align)

    ws2.cell(row=current_row + 1, column=2, value=desc)
    ws2.cell(row=current_row + 1, column=2).font = copy.copy(desc_font)
    ws2.cell(row=current_row + 1, column=2).alignment = copy.copy(c_align)

    img = XlImage(f"doc/{filename}")
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    ws2.add_image(img, f"B{current_row + 3}")

    current_row += ROW_GAP

wb.save("doc/화면이동테스트결과서.xlsx")
print("회원가입 테스트 케이스 + 캡처 시트 추가 완료")
