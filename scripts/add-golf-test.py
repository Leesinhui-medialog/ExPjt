# -*- coding: utf-8 -*-
"""
화면이동테스트결과서에 골프장 테스트 케이스 추가 + '골프장 캡처' 시트 생성.
"""
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("doc/화면이동테스트결과서.xlsx")

# ===== 1. 테스트 케이스 추가 =====
ws = wb["화면이동테스트결과서"]

data_font = Font(name="맑은 고딕", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

next_row = ws.max_row + 1

new_cases = [
    {
        "id": "SCT-F-GOLF-001", "type": "기능", "section": "화면이동",
        "major": "골프장", "minor": "골프장 화면 표시",
        "reqId": "", "funcId": "GOLF-SCR-001", "progId": "LocationDisplay.jsx", "screenId": "/location/display",
        "purpose": "골프장 화면 정상 표시 검증 (지도 + 검색 패널 + 골프장 마커)",
        "procedure": "1. 로그인 후 좌측 메뉴에서 '위치 정보' 클릭\n2. 골프장 화면(/location/display) 이동\n3. 지도, 검색 패널, 골프장 목록 표시 확인",
        "detail": "지도 위에 골프장 마커가 표시되고, 좌측 검색 패널에 전체 골프장 목록이 정상 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-GOLF-002", "type": "기능", "section": "화면이동",
        "major": "골프장", "minor": "골프장 검색",
        "reqId": "", "funcId": "GOLF-SCR-002", "progId": "SearchPanel.jsx", "screenId": "/location/display",
        "purpose": "골프장 키워드 검색 기능 검증 ('남서울' 검색)",
        "procedure": "1. 골프장 화면에서 검색 입력란에 '남서울' 입력\n2. 디바운스(300ms) 후 검색 결과 표시 확인\n3. 검색 결과 목록에 '남서울' 포함 골프장만 표시되는지 확인",
        "detail": "'남서울' 키워드 검색 시 해당 골프장이 검색 결과에 정상 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

for tc in new_cases:
    col_map = {
        2: tc["id"], 3: tc["type"], 4: tc["section"],
        5: tc["major"], 6: tc["minor"], 7: tc["reqId"],
        8: tc["funcId"], 9: tc["progId"], 10: tc["screenId"],
        11: tc["purpose"], 12: tc["procedure"], 13: tc["detail"],
        14: tc["date"], 15: tc["tester"], 16: tc["result"], 17: tc["reviewer"],
    }
    for col, value in col_map.items():
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)
    next_row += 1

# ===== 2. '골프장 캡처' 시트 생성 =====
if "골프장 캡처" in wb.sheetnames:
    del wb["골프장 캡처"]

ws2 = wb.create_sheet("골프장 캡처")

title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
c_align = Alignment(horizontal="center", vertical="center")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 100

# 골프장 전체 화면
ws2["B2"] = "1. 골프장 화면 — 전체 목록 + 지도"
ws2["B2"].font = title_font
ws2["B2"].alignment = c_align

ws2["B3"] = "로그인 후 /location/display 이동 시 지도와 골프장 목록이 정상 표시됨"
ws2["B3"].font = desc_font
ws2["B3"].alignment = c_align

img1 = Image("doc/07_골프장화면.png")
img1.width = 700
img1.height = int(img1.height * (700 / img1.width)) if img1.width > 0 else 500
ws2.add_image(img1, "B5")

# 검색 결과
start_row2 = 38
ws2.cell(row=start_row2, column=2, value="2. 골프장 검색 결과 — '남서울' 검색")
ws2.cell(row=start_row2, column=2).font = title_font
ws2.cell(row=start_row2, column=2).alignment = c_align

ws2.cell(row=start_row2 + 1, column=2, value="'남서울' 키워드 입력 후 해당 골프장이 검색 결과에 정상 표시됨")
ws2.cell(row=start_row2 + 1, column=2).font = desc_font
ws2.cell(row=start_row2 + 1, column=2).alignment = c_align

img2 = Image("doc/08_골프장검색결과.png")
img2.width = 700
img2.height = int(img2.height * (700 / img2.width)) if img2.width > 0 else 500
ws2.add_image(img2, f"B{start_row2 + 3}")

wb.save("doc/화면이동테스트결과서.xlsx")
print("골프장 테스트 케이스 + 캡처 시트 추가 완료")
