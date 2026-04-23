# -*- coding: utf-8 -*-
"""
화면 이동 테스트 결과서 엑셀 생성 스크립트.
doc/단위테스트결과서.xlsx 양식을 기반으로 새 파일(doc/화면이동테스트결과서.xlsx)을 생성한다.
"""
import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ===== 겉표지 시트 =====
ws_cover = wb.active
ws_cover.title = "겉표지"
ws_cover["L2"] = "ExPjt 샘플 애플리케이션"
ws_cover["L4"] = "화면이동 테스트 결과서"
ws_cover["L8"] = "문서번호"
ws_cover["L9"] = "Var. 1.0"
ws_cover["A20"] = "Copyright (C) 미디어로그"
ws_cover["A21"] = "미디어로그의 사전승인 없이 본 내용의 전부 또는 일부에 대한 복제, 전재, 배포, 사용을 금합니다."

# ===== 계정이력 시트 =====
ws_history = wb.create_sheet("계정이력")
ws_history["A1"] = "=겉표지!L2"
ws_history["E1"] = "=겉표지!L4"
ws_history["A3"] = "계 정 이 력"
headers_hist = ["버전", "변경이력", "변경내용", "작성자", "승인자"]
for i, h in enumerate(headers_hist, 1):
    ws_history.cell(row=6, column=i, value=h)
ws_history["A7"] = "V1.0"
ws_history["B7"] = "2026-04-20"
ws_history["C7"] = "최초작성"
ws_history["D7"] = "sinhuiyo"

# ===== 화면이동테스트결과서 시트 =====
ws = wb.create_sheet("화면이동테스트결과서")

# 스타일 정의
header_font = Font(name="맑은 고딕", size=10, bold=True)
header_fill = PatternFill("solid", fgColor="FFFF00")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
data_font = Font(name="맑은 고딕", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 헤더 (2행, B~Q열)
headers = [
    "시험ID", "시험유형", "시험구분", "시험항목(대분류)", "시험항목(중분류)",
    "요구사항ID", "기능ID", "프로그램ID", "화면ID", "시험목적",
    "시험절차", "시험내역", "시험일자\n(yyyy-mm-dd)", "시험자", "시험결과",
    "시험결과 확인자\n(개발PL/PM)"
]
for col_idx, header in enumerate(headers, 2):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = copy.copy(header_font)
    cell.fill = copy.copy(header_fill)
    cell.alignment = copy.copy(header_align)
    cell.border = copy.copy(thin_border)

# 열 너비 설정
col_widths = {
    "B": 16, "C": 10, "D": 12, "E": 16, "F": 16,
    "G": 12, "H": 12, "I": 14, "J": 12, "K": 30,
    "L": 40, "M": 35, "N": 18, "O": 10, "P": 10, "Q": 18
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 행 높이
ws.row_dimensions[2].height = 25.5

# 테스트 케이스 데이터
test_cases = [
    {
        "id": "SCT-F-LGIN-001", "type": "기능", "section": "화면이동",
        "major": "로그인", "minor": "로그인 화면 접속",
        "reqId": "", "funcId": "LGIN-SCR-001", "progId": "Login.jsx", "screenId": "/login",
        "purpose": "http://localhost:3000/ 접속 시 로그인 화면 정상 표시 검증",
        "procedure": "1. 브라우저에서 http://localhost:3000/ 접속\n2. 로그인 화면이 표시되는지 확인",
        "detail": "로그인 폼(이메일, 비밀번호 입력란, 로그인 버튼)이 정상적으로 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-LGIN-002", "type": "기능", "section": "화면이동",
        "major": "로그인", "minor": "로그인 처리",
        "reqId": "", "funcId": "LGIN-SCR-002", "progId": "Login.jsx", "screenId": "/login",
        "purpose": "admin@medialog.co.kr / admin123 계정으로 로그인 성공 검증",
        "procedure": "1. 이메일: admin@medialog.co.kr 입력\n2. 비밀번호: admin123 입력\n3. 로그인 버튼 클릭",
        "detail": "환영 메시지 팝업 표시 후 게시판 목록 화면(/board/list)으로 자동 이동",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-BORD-001", "type": "기능", "section": "화면이동",
        "major": "게시판", "minor": "게시판 목록 화면",
        "reqId": "", "funcId": "BORD-SCR-001", "progId": "BoardList.jsx", "screenId": "/board/list",
        "purpose": "로그인 후 게시판 목록 화면 정상 표시 검증",
        "procedure": "1. 로그인 성공 후 자동 이동된 게시판 목록 화면 확인\n2. 좌측 메뉴, 상단 헤더, 게시글 목록 테이블 표시 확인",
        "detail": "게시판 목록 화면이 정상 표시됨 (좌측 메뉴, 로그아웃 버튼, 게시글 테이블 포함)",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

# 데이터 행 작성
for idx, tc in enumerate(test_cases):
    row = idx + 3
    col_map = {
        2: tc["id"], 3: tc["type"], 4: tc["section"],
        5: tc["major"], 6: tc["minor"], 7: tc["reqId"],
        8: tc["funcId"], 9: tc["progId"], 10: tc["screenId"],
        11: tc["purpose"], 12: tc["procedure"], 13: tc["detail"],
        14: tc["date"], 15: tc["tester"], 16: tc["result"], 17: tc["reviewer"],
    }
    for col, value in col_map.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)

# 저장
output_path = "doc/화면이동테스트결과서.xlsx"
wb.save(output_path)
print(f"총 {len(test_cases)}건의 테스트 케이스를 작성했습니다.")
print(f"{output_path} 저장 완료")
