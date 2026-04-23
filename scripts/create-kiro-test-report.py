# -*- coding: utf-8 -*-
"""
Kiro테스트결과서_20260420.xlsx 생성 스크립트.
기존 Kiro테스트결과서.xlsx의 단위테스트결과서 시트를 참고하여:
1. 시험일자=오늘, 시험자=sinhuiyo, 시험결과=Pass/Fail 채움
2. 대분류별 캡처 시트 생성 (이미지 겹치지 않게)
3. R열에 에러 사유 (Fail인 경우)
4. 계정이력 버전 0.1 올림
5. 테스트주의사항 시트에 규칙 추가
"""
import copy, json, os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TODAY = "2026-04-20"
TESTER = "sinhuiyo"
IMG_DIR = "doc/KiroImg"

# 기존 파일 로드
wb = load_workbook("doc/Kiro테스트결과서_20260420.xlsx")

# 결과 JSON 로드
with open(os.path.join(IMG_DIR, "results.json"), "r", encoding="utf-8") as f:
    results = json.load(f)
result_map = {r["id"]: r for r in results}

# 스타일
data_font = Font(name="맑은 고딕", size=10)
fail_fill = PatternFill("solid", fgColor="FFCCCC")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
c_align = Alignment(horizontal="center", vertical="center")

# ===== 1. 단위테스트결과서 시트에 시험일자/시험자/시험결과/에러사유 채우기 =====
ws = wb["단위테스트결과서"]

# R열 헤더 추가 (에러사유)
r_header = ws.cell(row=2, column=18, value="에러사유")
r_header.font = Font(name="맑은 고딕", size=10, bold=True)
r_header.fill = PatternFill("solid", fgColor="FFFF00")
r_header.alignment = copy.copy(center_align)
r_header.border = copy.copy(thin_border)

for row in range(3, ws.max_row + 1):
    test_id = ws.cell(row=row, column=2).value  # B열: 시험ID
    if not test_id:
        continue

    # N열: 시험일자
    cell_n = ws.cell(row=row, column=14, value=TODAY)
    cell_n.font = copy.copy(data_font)
    cell_n.alignment = copy.copy(center_align)
    cell_n.border = copy.copy(thin_border)

    # O열: 시험자
    cell_o = ws.cell(row=row, column=15, value=TESTER)
    cell_o.font = copy.copy(data_font)
    cell_o.alignment = copy.copy(center_align)
    cell_o.border = copy.copy(thin_border)

    # P열: 시험결과
    r = result_map.get(test_id)
    result_val = "Pass"
    error_msg = ""
    if r:
        result_val = r["result"]
        error_msg = r.get("error", "")
        # 캡처 이미지 파일명에서 실패 키워드 감지
        img_file = r.get("img", "")
        # 에러 메시지가 있으면 무조건 Fail
        if error_msg:
            result_val = "Fail"
        # result가 이미 Fail이면 유지
        if r["result"] == "Fail":
            result_val = "Fail"

    cell_p = ws.cell(row=row, column=16, value=result_val)
    cell_p.font = copy.copy(data_font)
    cell_p.alignment = copy.copy(center_align)
    cell_p.border = copy.copy(thin_border)
    if result_val == "Fail":
        cell_p.fill = copy.copy(fail_fill)
        # 행 전체에 연한 빨간 배경 적용
        for c in range(2, 18):
            ws.cell(row=row, column=c).fill = copy.copy(fail_fill)

    # R열: 에러사유
    cell_r = ws.cell(row=row, column=18, value=error_msg)
    cell_r.font = copy.copy(data_font)
    cell_r.alignment = copy.copy(left_align)
    cell_r.border = copy.copy(thin_border)
    if error_msg:
        cell_r.fill = copy.copy(fail_fill)

ws.column_dimensions["R"].width = 40

# ===== 2. 대분류별 캡처 시트 생성 =====
# 대분류별 이미지 매핑
major_images = {
    "로그인": [
        ("LGIN_01_로그인화면.png", "로그인 화면 접속", "http://localhost:3000/ 접속 시 로그인 폼 정상 표시"),
        ("LGIN_02_로그인성공.png", "로그인 처리", "admin@medialog.co.kr 로그인 성공 후 게시판 목록 이동"),
    ],
    "게시판": [
        ("BORD_01_게시판목록.png", "게시판 목록 화면", "로그인 후 게시판 목록 정상 표시"),
        ("BORD_02_글쓰기_저장전.png", "게시글 작성 화면", "제목/내용/첨부파일 입력 완료 상태"),
        ("BORD_03_저장결과.png", "게시글 저장 결과", "저장 후 게시판 목록에 신규 게시글 표시"),
        ("BORD_04_상세화면.png", "게시글 상세 화면", "제목 클릭 시 상세 화면 정상 표시"),
        ("BORD_05_첨부파일다운로드.png", "첨부파일 다운로드", "첨부파일 링크 클릭 후 다운로드 정상 동작"),
    ],
    "골프장": [
        ("GOLF_01_골프장화면.png", "골프장 화면 표시", "지도 + 검색 패널 + 골프장 마커 정상 표시"),
        ("GOLF_02_검색결과.png", "골프장 검색", "'남서울' 키워드 검색 결과 정상 표시"),
    ],
    "회원": [
        ("MEMB_01_Step01_약관동의.png", "Step 01 약관동의", "약관동의 화면 초기 상태"),
        ("MEMB_02_Step01_동의완료.png", "Step 01 약관동의 완료", "이용약관, 개인정보 수집 동의 체크 완료"),
        ("MEMB_03_Step02_실명확인.png", "Step 02 실명확인", "이름/생년월일 입력 화면"),
        ("MEMB_04_Step02_인증완료.png", "Step 02 실명인증 완료", "실명인증 완료 상태"),
        ("MEMB_05_Step03_정보입력.png", "Step 03 회원정보입력", "이메일/전화번호/비밀번호 입력 화면"),
        ("MEMB_06_Step03_입력완료.png", "Step 03 정보입력 완료", "모든 회원정보 입력 완료 상태"),
        ("MEMB_07_가입완료.png", "가입완료", "회원가입 완료 팝업 표시"),
    ],
}

IMG_WIDTH = 700
IMG_HEIGHT = 492
ROW_GAP = 37  # 이미지 높이(33행) + 제목(2행) + 여백(2행)

for major_name, img_list in major_images.items():
    sheet_name = f"{major_name} 캡처"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws2 = wb.create_sheet(sheet_name)
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 100

    current_row = 2
    for idx, (filename, subtitle, description) in enumerate(img_list):
        img_path = os.path.join(IMG_DIR, filename)
        if not os.path.exists(img_path):
            continue

        # 제목
        ws2.cell(row=current_row, column=2, value=f"{idx+1}. {subtitle}")
        ws2.cell(row=current_row, column=2).font = copy.copy(title_font)
        ws2.cell(row=current_row, column=2).alignment = copy.copy(c_align)

        # 설명
        ws2.cell(row=current_row + 1, column=2, value=description)
        ws2.cell(row=current_row + 1, column=2).font = copy.copy(desc_font)
        ws2.cell(row=current_row + 1, column=2).alignment = copy.copy(c_align)

        # 이미지
        img = XlImage(img_path)
        img.width = IMG_WIDTH
        img.height = IMG_HEIGHT
        ws2.add_image(img, f"B{current_row + 3}")

        current_row += ROW_GAP

# ===== 3. 계정이력 업데이트 =====
ws_hist = wb["계정이력"]
# 기존 마지막 버전 확인
last_ver_row = 7
for r in range(7, ws_hist.max_row + 1):
    if ws_hist.cell(row=r, column=1).value:
        last_ver_row = r

new_row = last_ver_row + 1
hist_data = {
    1: "V1.1",
    2: TODAY,
    3: "Kiro 사용자테스트",
    4: TESTER,
    5: "",
}
for col, value in hist_data.items():
    cell = ws_hist.cell(row=new_row, column=col, value=value)
    cell.font = copy.copy(data_font)
    cell.alignment = copy.copy(center_align)
    cell.border = copy.copy(thin_border)

# 겉표지 버전 업데이트
wb["겉표지"]["L9"] = "Var. 1.1"

# ===== 4. 테스트주의사항 시트 업데이트 =====
ws_note = wb["테스트주의사항"]

notes = [
    "1. 시험일자는 오늘 일자로 기입한다.",
    "2. 시험자는 Kiro 로그인 계정(sinhuiyo)을 기입한다.",
    "3. 시험결과는 정상인 경우 Pass, 에러 발생 시 Fail로 기입한다.",
    "4. 각 기능 테스트 시 시험항목(대분류) 단위로 시트를 만들고 기능테스트 결과를 캡처하여 차례대로 넣는다.",
    "5. 캡처된 화면은 겹치지 않게 이미지를 정리하고, 시험항목(중분류) 단위로 타이틀을 넣는다.",
    "6. 에러가 있는 경우 Fail로 마킹하고, 에러 사유를 R열에 기입한다.",
    "7. 테스트 완료 시 계정이력에 버전을 0.1 올리고, 변경일은 오늘, 변경내용은 'Kiro 사용자테스트', 작성자는 Kiro 로그인 계정을 기입한다.",
]

note_font = Font(name="맑은 고딕", size=11)
note_align = Alignment(vertical="center", wrap_text=True)

# 기존 내용 아래에 추가
start_row = 4
for i, note in enumerate(notes):
    cell = ws_note.cell(row=start_row + i, column=2, value=note)
    cell.font = copy.copy(note_font)
    cell.alignment = copy.copy(note_align)

ws_note.column_dimensions["B"].width = 100

# 저장
output_path = f"doc/Kiro테스트결과서_{TODAY.replace('-', '')}_v2.xlsx"
wb.save(output_path)
print(f"파일 생성 완료: {output_path}")
