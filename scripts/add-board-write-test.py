# -*- coding: utf-8 -*-
"""
화면이동테스트결과서에 게시글 작성 테스트 케이스 추가 + 캡처화면 시트 추가.
"""
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("doc/화면이동테스트결과서.xlsx")

# ===== 1. 테스트 케이스 시트에 게시글 작성 케이스 추가 =====
ws = wb["화면이동테스트결과서"]

data_font = Font(name="맑은 고딕", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 기존 데이터 행 수 확인 (2행=헤더, 3행부터 데이터)
next_row = ws.max_row + 1

new_cases = [
    {
        "id": "SCT-F-BORD-002", "type": "기능", "section": "화면이동",
        "major": "게시판", "minor": "게시글 작성 화면",
        "reqId": "", "funcId": "BORD-SCR-002", "progId": "BoardWrite.jsx", "screenId": "/board/write",
        "purpose": "게시글 작성 화면에서 제목/내용/첨부파일 입력 후 저장 전 화면 검증",
        "procedure": "1. 로그인 후 글쓰기 페이지 이동\n2. 제목: '게시물 작성 테스트 AI' 입력\n3. 첨부파일: mysql.png 업로드\n4. 내용 입력\n5. 저장 전 화면 확인",
        "detail": "제목, 첨부파일명, 내용이 모두 정상 입력된 상태로 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-BORD-003", "type": "기능", "section": "화면이동",
        "major": "게시판", "minor": "게시글 저장 결과",
        "reqId": "", "funcId": "BORD-SCR-003", "progId": "BoardWrite.jsx", "screenId": "/board/list",
        "purpose": "게시글 저장 후 게시판 목록에 신규 게시글 표시 검증",
        "procedure": "1. 저장 버튼 클릭\n2. 게시판 목록 화면으로 자동 이동\n3. 작성한 게시글이 목록에 표시되는지 확인",
        "detail": "'게시물 작성 테스트 AI' 제목의 게시글이 게시판 목록에 정상 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

for tc in new_cases:
    col_map = {
        2: tc["id"], 3: tc["type"], 4: tc["section"],
        5: tc["major"], 6: tc["minor"], 7: tc["reqId"],
        8: tc["funcId"], 9: tc["progId"], 10: tc["screenId"],
        11: tc["purpose"], 12: tc["procedure"], 13: tc["detail"],
        14: tc["date"], 15: tc["tester"], 16: tc["result"], 17: tc["reviewer"],
    }
    for col, value in col_map.items():
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)
    next_row += 1

# ===== 2. 게시글작성 캡처화면 시트 추가 =====
if "게시글작성 캡처" in wb.sheetnames:
    del wb["게시글작성 캡처"]

ws2 = wb.create_sheet("게시글작성 캡처")

title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
c_align = Alignment(horizontal="center", vertical="center")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 100

# 저장 전 화면
ws2["B2"] = "1. 게시글 작성 — 저장 전 화면"
ws2["B2"].font = title_font
ws2["B2"].alignment = c_align

ws2["B3"] = "제목: '게시물 작성 테스트 AI' / 첨부파일: mysql.png / 내용 입력 완료 상태"
ws2["B3"].font = desc_font
ws2["B3"].alignment = c_align

img1 = Image("doc/03_게시글작성_저장전.png")
img1.width = 700
img1.height = int(img1.height * (700 / img1.width)) if img1.width > 0 else 500
ws2.add_image(img1, "B5")

# 저장 결과 화면
start_row2 = 38
ws2.cell(row=start_row2, column=2, value="2. 게시글 작성 — 저장 결과 화면")
ws2.cell(row=start_row2, column=2).font = title_font
ws2.cell(row=start_row2, column=2).alignment = c_align

ws2.cell(row=start_row2 + 1, column=2, value="저장 후 게시판 목록에 '게시물 작성 테스트 AI' 게시글이 정상 표시됨")
ws2.cell(row=start_row2 + 1, column=2).font = desc_font
ws2.cell(row=start_row2 + 1, column=2).alignment = c_align

img2 = Image("doc/04_게시글작성_저장결과.png")
img2.width = 700
img2.height = int(img2.height * (700 / img2.width)) if img2.width > 0 else 500
ws2.add_image(img2, f"B{start_row2 + 3}")

wb.save("doc/화면이동테스트결과서.xlsx")
print("게시글 작성 테스트 케이스 + 캡처화면 시트 추가 완료")
