# -*- coding: utf-8 -*-
"""
화면이동테스트결과서에 '캡처화면' 시트를 추가하고 스크린샷 이미지를 삽입한다.
"""
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

wb = load_workbook("doc/화면이동테스트결과서.xlsx")

# 기존에 '캡처화면' 시트가 있으면 삭제 후 재생성
if "캡처화면" in wb.sheetnames:
    del wb["캡처화면"]

ws = wb.create_sheet("캡처화면")

# 제목 스타일
title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
center_align = Alignment(horizontal="center", vertical="center")

# 열 너비 설정
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 100

# 1. 로그인 화면
ws["B2"] = "1. 로그인 화면 (http://localhost:3000/)"
ws["B2"].font = title_font
ws["B2"].alignment = center_align

ws["B3"] = "접속 시 로그인 폼(이메일, 비밀번호, 로그인 버튼)이 정상 표시됨"
ws["B3"].font = desc_font
ws["B3"].alignment = center_align

img1 = Image("doc/01_로그인화면.png")
img1.width = 700
img1.height = int(img1.height * (700 / img1.width)) if img1.width > 0 else 500
ws.add_image(img1, "B5")

# 이미지 높이에 맞춰 행 간격 확보 (대략 30행 정도)
start_row2 = 35

# 2. 게시판 목록 화면
ws.cell(row=start_row2, column=2, value="2. 게시판 목록 화면 (로그인 후 /board/list)")
ws.cell(row=start_row2, column=2).font = title_font
ws.cell(row=start_row2, column=2).alignment = center_align

ws.cell(row=start_row2 + 1, column=2, value="admin@medialog.co.kr / admin123 로그인 후 게시판 목록 화면이 정상 표시됨")
ws.cell(row=start_row2 + 1, column=2).font = desc_font
ws.cell(row=start_row2 + 1, column=2).alignment = center_align

img2 = Image("doc/02_게시판목록화면.png")
img2.width = 700
img2.height = int(img2.height * (700 / img2.width)) if img2.width > 0 else 500
ws.add_image(img2, f"B{start_row2 + 3}")

wb.save("doc/화면이동테스트결과서.xlsx")
print("캡처화면 시트 추가 완료 — doc/화면이동테스트결과서.xlsx")
