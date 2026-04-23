# -*- coding: utf-8 -*-
"""
화면이동테스트결과서에 게시글 상세보기/첨부파일 다운로드 테스트 케이스 추가
+ '게시글작성 캡처' 시트에 캡처 이미지 추가.
"""
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("doc/화면이동테스트결과서.xlsx")

# ===== 1. 테스트 케이스 추가 =====
ws = wb["화면이동테스트결과서"]

data_font = Font(name="맑은 고딕", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

next_row = ws.max_row + 1

new_cases = [
    {
        "id": "SCT-F-BORD-004", "type": "기능", "section": "화면이동",
        "major": "게시판", "minor": "게시글 상세 화면",
        "reqId": "", "funcId": "BORD-SCR-004", "progId": "BoardDetail.jsx", "screenId": "/board/detail/{idx}",
        "purpose": "게시판 목록에서 제목 클릭 시 상세 화면 정상 표시 검증",
        "procedure": "1. 게시판 목록에서 게시글 제목 클릭\n2. 상세 화면으로 이동\n3. 제목, 내용, 첨부파일, 작성자, 작성일 표시 확인",
        "detail": "게시글 상세 화면에 제목, 내용, 첨부파일(mysql.png), 작성자, 작성일, 조회수가 정상 표시됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "SCT-F-BORD-005", "type": "기능", "section": "화면이동",
        "major": "게시판", "minor": "첨부파일 다운로드",
        "reqId": "", "funcId": "BORD-SCR-005", "progId": "BoardDetail.jsx", "screenId": "/board/detail/{idx}",
        "purpose": "게시글 상세 화면에서 첨부파일 다운로드 기능 검증",
        "procedure": "1. 게시글 상세 화면에서 첨부파일 링크 클릭\n2. 파일 다운로드 실행\n3. 다운로드된 파일 확인",
        "detail": "첨부파일(mysql.png)이 정상적으로 다운로드됨",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

for tc in new_cases:
    col_map = {
        2: tc["id"], 3: tc["type"], 4: tc["section"],
        5: tc["major"], 6: tc["minor"], 7: tc["reqId"],
        8: tc["funcId"], 9: tc["progId"], 10: tc["screenId"],
        11: tc["purpose"], 12: tc["procedure"], 13: tc["detail"],
        14: tc["date"], 15: tc["tester"], 16: tc["result"], 17: tc["reviewer"],
    }
    for col, value in col_map.items():
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)
    next_row += 1

# ===== 2. '게시글작성 캡처' 시트에 상세보기/다운로드 캡처 추가 =====
ws2 = wb["게시글작성 캡처"]

title_font = Font(name="맑은 고딕", size=14, bold=True)
desc_font = Font(name="맑은 고딕", size=11)
c_align = Alignment(horizontal="center", vertical="center")

# 기존 이미지 아래에 추가 (기존 2개 이미지 이후 = 약 70행 이후)
start_row = 72

# 상세 화면
ws2.cell(row=start_row, column=2, value="3. 게시글 상세 화면 — 목록에서 제목 클릭 후")
ws2.cell(row=start_row, column=2).font = title_font
ws2.cell(row=start_row, column=2).alignment = c_align

ws2.cell(row=start_row + 1, column=2, value="게시글 제목, 내용, 첨부파일(mysql.png), 작성자, 작성일, 조회수가 정상 표시됨")
ws2.cell(row=start_row + 1, column=2).font = desc_font
ws2.cell(row=start_row + 1, column=2).alignment = c_align

img1 = Image("doc/05_게시글상세화면.png")
img1.width = 700
img1.height = int(img1.height * (700 / img1.width)) if img1.width > 0 else 500
ws2.add_image(img1, f"B{start_row + 3}")

# 첨부파일 다운로드
start_row2 = start_row + 38
ws2.cell(row=start_row2, column=2, value="4. 첨부파일 다운로드 — mysql.png 다운로드 완료")
ws2.cell(row=start_row2, column=2).font = title_font
ws2.cell(row=start_row2, column=2).alignment = c_align

ws2.cell(row=start_row2 + 1, column=2, value="첨부파일 링크 클릭 후 mysql.png 파일이 정상적으로 다운로드됨")
ws2.cell(row=start_row2 + 1, column=2).font = desc_font
ws2.cell(row=start_row2 + 1, column=2).alignment = c_align

img2 = Image("doc/06_첨부파일다운로드.png")
img2.width = 700
img2.height = int(img2.height * (700 / img2.width)) if img2.width > 0 else 500
ws2.add_image(img2, f"B{start_row2 + 3}")

wb.save("doc/화면이동테스트결과서.xlsx")
print("게시글 상세보기/다운로드 테스트 케이스 + 캡처 추가 완료")
