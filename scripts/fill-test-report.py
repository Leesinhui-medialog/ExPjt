# -*- coding: utf-8 -*-
"""
단위테스트결과서 엑셀 자동 작성 스크립트.
기존 doc/단위테스트결과서.xlsx 템플릿에 프로젝트 전체 테스트 케이스를 채워넣는다.
"""
import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 기존 엑셀 로드
wb = load_workbook("doc/단위테스트결과서.xlsx")
ws = wb["단위테스트결과서"]

# 3행(샘플 데이터) 스타일을 복사하여 데이터 행에 적용할 기본 스타일 정의
data_font = Font(name="LG스마트체 Regular", size=10, color="000000")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 테스트 케이스 데이터 정의
# 컬럼: B=시험ID, C=시험유형, D=시험구분, E=시험항목(대분류), F=시험항목(중분류),
#        G=요구사항ID, H=기능ID, I=프로그램ID, J=화면ID, K=시험목적,
#        L=시험절차, M=시험내역, N=시험일자, O=시험자, P=시험결과, Q=시험결과확인자
test_cases = [
    # ===== 메인(MAIN) =====
    {
        "id": "UST-F-MAIN-001", "type": "기능", "section": "단위",
        "major": "메인", "minor": "HelloWorld",
        "reqId": "", "funcId": "MAIN-001", "progId": "HelloWorldServiceTest", "screenId": "",
        "purpose": "HelloWorld 메시지 반환 검증",
        "procedure": "1. HelloWorldService.retrieve() 호출",
        "detail": "반환값이 'Hello, World!'인지 assertEquals로 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIN-002", "type": "기능", "section": "단위",
        "major": "메인", "minor": "HelloWorld",
        "reqId": "", "funcId": "MAIN-002", "progId": "HelloWorldControllerTest", "screenId": "",
        "purpose": "GET /hello 엔드포인트 동작 검증",
        "procedure": "1. MockMvc로 GET /hello 요청\n2. 응답 상태 및 본문 확인",
        "detail": "status 200, 본문 'Hello, World!' 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 로그인(LGIN) =====
    {
        "id": "UST-F-LGIN-001", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 인증",
        "reqId": "", "funcId": "LGIN-001", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "로그인 성공 검증",
        "procedure": "1. 유효한 이메일/비밀번호로 authenticate 호출",
        "detail": "success=true, memberName='홍길동' 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-002", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 인증",
        "reqId": "", "funcId": "LGIN-002", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "존재하지 않는 이메일 로그인 실패 검증",
        "procedure": "1. 미등록 이메일로 authenticate 호출",
        "detail": "success=false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-003", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 인증",
        "reqId": "", "funcId": "LGIN-003", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "비밀번호 불일치 로그인 실패 검증",
        "procedure": "1. 올바른 이메일 + 잘못된 비밀번호로 authenticate 호출",
        "detail": "success=false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-004", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 인증",
        "reqId": "", "funcId": "LGIN-004", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "삭제된 회원 로그인 실패 검증",
        "procedure": "1. delYn='Y'인 회원으로 authenticate 호출",
        "detail": "success=false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-005", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "비밀번호 찾기",
        "reqId": "", "funcId": "LGIN-005", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "비밀번호 찾기 성공 — 임시비밀번호 발급 및 메일 발송",
        "procedure": "1. 등록된 이메일로 retrievePassword 호출",
        "detail": "success=true, 비밀번호 변경 확인, 메일 발송 verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-006", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "비밀번호 찾기",
        "reqId": "", "funcId": "LGIN-006", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "비밀번호 찾기 실패 — 미등록 이메일",
        "procedure": "1. 미등록 이메일로 retrievePassword 호출",
        "detail": "success=false, 메일 미발송 verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-007", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "비밀번호 찾기",
        "reqId": "", "funcId": "LGIN-007", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "비밀번호 찾기 실패 — 삭제된 회원",
        "procedure": "1. delYn='Y'인 회원 이메일로 retrievePassword 호출",
        "detail": "success=false, 메일 미발송 verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-008", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "비밀번호 찾기",
        "reqId": "", "funcId": "LGIN-008", "progId": "LoginServiceTest", "screenId": "",
        "purpose": "비밀번호 찾기 — 메일 발송 실패해도 성공 응답",
        "procedure": "1. 메일 발송 예외 설정 후 retrievePassword 호출",
        "detail": "success=true 반환 검증 (메일 실패 무시)",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-009", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-009", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "로그인 API 성공 시 세션 저장 검증",
        "procedure": "1. POST /api/login 유효 자격증명 요청\n2. 세션 속성 확인",
        "detail": "status 200, 세션에 loginEmail/loginMemberName 저장 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-010", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-010", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "로그인 API 실패 시 세션 미저장 검증",
        "procedure": "1. POST /api/login 잘못된 비밀번호 요청\n2. 세션 속성 확인",
        "detail": "success=false, 세션에 loginEmail 미저장 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-011", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-011", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "로그인 상태 확인 API — 로그인 상태",
        "procedure": "1. 세션에 로그인 정보 설정\n2. GET /api/login/check 요청",
        "detail": "loggedIn=true, email/memberName 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-012", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-012", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "로그인 상태 확인 API — 비로그인 상태",
        "procedure": "1. 빈 세션으로 GET /api/login/check 요청",
        "detail": "loggedIn=false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-013", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-013", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "로그아웃 API — 세션 무효화 검증",
        "procedure": "1. 로그인 세션 설정\n2. POST /api/login/logout 요청",
        "detail": "result='ok', 세션 무효화(isInvalid) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-014", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-014", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "비밀번호 찾기 API — 성공",
        "procedure": "1. POST /api/login/retrieve-password 등록 이메일 요청",
        "detail": "success=true, 임시비밀번호 발송 메시지 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-LGIN-015", "type": "기능", "section": "서비스구간",
        "major": "로그인", "minor": "로그인 컨트롤러",
        "reqId": "", "funcId": "LGIN-015", "progId": "LoginControllerTest", "screenId": "",
        "purpose": "비밀번호 찾기 API — 실패 (회원 없음)",
        "procedure": "1. POST /api/login/retrieve-password 미등록 이메일 요청",
        "detail": "success=false, '회원 정보가 존재하지 않습니다.' 메시지 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 게시판(BORD) =====
    {
        "id": "UST-F-BORD-001", "type": "기능", "section": "단위",
        "major": "게시판", "minor": "게시판 리포지토리",
        "reqId": "", "funcId": "BORD-001", "progId": "BoardRepositoryTest", "screenId": "",
        "purpose": "저장 후 전체 조회 검증",
        "procedure": "1. Board 엔티티 저장\n2. findAll 호출",
        "detail": "저장된 게시글 1건 조회, 제목 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-002", "type": "기능", "section": "단위",
        "major": "게시판", "minor": "게시판 리포지토리",
        "reqId": "", "funcId": "BORD-002", "progId": "BoardRepositoryTest", "screenId": "",
        "purpose": "ID로 조회 검증",
        "procedure": "1. Board 저장\n2. findById 호출",
        "detail": "조회 결과 not null, 제목 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-003", "type": "기능", "section": "단위",
        "major": "게시판", "minor": "게시판 리포지토리",
        "reqId": "", "funcId": "BORD-003", "progId": "BoardRepositoryTest", "screenId": "",
        "purpose": "삭제 후 조회 시 빈 결과 검증",
        "procedure": "1. Board 저장\n2. deleteById 호출\n3. findById 확인",
        "detail": "삭제 후 findById 결과 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-004", "type": "기능", "section": "단위",
        "major": "게시판", "minor": "게시판 리포지토리",
        "reqId": "", "funcId": "BORD-004", "progId": "BoardRepositoryTest", "screenId": "",
        "purpose": "작성자 이름(authorName) 저장 및 조회 검증",
        "procedure": "1. authorName 설정 후 저장\n2. findById로 조회",
        "detail": "authorName '홍길동' 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-005", "type": "기능", "section": "단위",
        "major": "게시판", "minor": "게시판 리포지토리",
        "reqId": "", "funcId": "BORD-005", "progId": "BoardRepositoryTest", "screenId": "",
        "purpose": "조회수(viewCount) 기본값 및 수정 검증",
        "procedure": "1. Board 저장 후 기본값 확인\n2. viewCount 수정 후 재조회",
        "detail": "기본값 0, 수정 후 5 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-006", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-006", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "전체 목록 조회 검증",
        "procedure": "1. boardRepository.findAll 모킹\n2. retrieveAll 호출",
        "detail": "반환 리스트 크기 1, 제목 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-007", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-007", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "페이징 조회 검증",
        "procedure": "1. findByDelYn 모킹\n2. retrievePage(0, 10) 호출",
        "detail": "totalElements 1 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-008", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-008", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "ID로 게시글 조회 성공 검증",
        "procedure": "1. findById 모킹(존재)\n2. retrieveById(1) 호출",
        "detail": "결과 not null, 제목 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-009", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-009", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "ID로 게시글 조회 실패 검증",
        "procedure": "1. findById 모킹(미존재)\n2. retrieveById(999) 호출",
        "detail": "결과 null 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-010", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-010", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "저장 시 등록일 자동 설정 검증",
        "procedure": "1. Board 저장 호출",
        "detail": "regDate not null, delYn='N' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-011", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-011", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "저장 시 메일 알림 서비스 호출 검증",
        "procedure": "1. Board 저장 호출\n2. mailNotificationService.sendNotification verify",
        "detail": "제목 '게시물 신규등록', 게시글 제목/내용 전달 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-012", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-012", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "메일 알림 실패 시에도 게시글 저장 정상 진행 검증",
        "procedure": "1. 메일 발송 예외 설정\n2. Board 저장 호출",
        "detail": "assertDoesNotThrow, boardRepository.save verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-013", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-013", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "소프트 삭제 시 delYn='Y' 변경 검증",
        "procedure": "1. softDelete(1) 호출",
        "detail": "delYn='Y', boardRepository.save verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-014", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 서비스",
        "reqId": "", "funcId": "BORD-014", "progId": "BoardServiceTest", "screenId": "",
        "purpose": "첨부파일 있는 게시글 삭제 시 물리 파일 삭제 검증",
        "procedure": "1. filePath 설정된 Board로 softDelete 호출",
        "detail": "delYn='Y', fileUploadService.deleteByPath verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-015", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 컨트롤러",
        "reqId": "", "funcId": "BORD-015", "progId": "BoardControllerTest", "screenId": "",
        "purpose": "목록 조회 시 JSON 배열 반환 검증",
        "procedure": "1. GET /api/board/list 요청",
        "detail": "status 200, content 배열 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-016", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 컨트롤러",
        "reqId": "", "funcId": "BORD-016", "progId": "BoardControllerTest", "screenId": "",
        "purpose": "상세 조회 시 게시글 정보 반환 검증",
        "procedure": "1. GET /api/board/1 요청",
        "detail": "status 200, title='테스트' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-BORD-017", "type": "기능", "section": "서비스구간",
        "major": "게시판", "minor": "게시판 컨트롤러",
        "reqId": "", "funcId": "BORD-017", "progId": "BoardControllerTest", "screenId": "",
        "purpose": "존재하지 않는 게시글 조회 시 404 반환 검증",
        "procedure": "1. GET /api/board/999 요청",
        "detail": "status 404 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 공통(COMM) — 파일 업로드 =====
    {
        "id": "UST-F-COMM-001", "type": "기능", "section": "단위",
        "major": "공통", "minor": "파일 업로드 리포지토리",
        "reqId": "", "funcId": "COMM-001", "progId": "UploadFileRepositoryTest", "screenId": "",
        "purpose": "저장 후 ID로 조회 검증",
        "procedure": "1. UploadFile 저장\n2. findById 호출",
        "detail": "originalName, fileSize 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-COMM-002", "type": "기능", "section": "단위",
        "major": "공통", "minor": "파일 업로드 리포지토리",
        "reqId": "", "funcId": "COMM-002", "progId": "UploadFileRepositoryTest", "screenId": "",
        "purpose": "삭제 후 조회 시 빈 결과 검증",
        "procedure": "1. UploadFile 저장\n2. deleteById 호출\n3. findById 확인",
        "detail": "삭제 후 findById 결과 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-COMM-003", "type": "기능", "section": "서비스구간",
        "major": "공통", "minor": "파일 업로드 서비스",
        "reqId": "", "funcId": "COMM-003", "progId": "FileUploadServiceTest", "screenId": "",
        "purpose": "파일 업로드 성공 검증",
        "procedure": "1. MockMultipartFile 생성\n2. upload('docs') 호출",
        "detail": "결과 not null, originalName 일치, save verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-COMM-004", "type": "기능", "section": "서비스구간",
        "major": "공통", "minor": "파일 업로드 서비스",
        "reqId": "", "funcId": "COMM-004", "progId": "FileUploadServiceTest", "screenId": "",
        "purpose": "빈 파일 업로드 시 예외 발생 검증",
        "procedure": "1. 빈 MockMultipartFile로 upload 호출",
        "detail": "IllegalArgumentException 발생 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-COMM-005", "type": "기능", "section": "서비스구간",
        "major": "공통", "minor": "파일 업로드 서비스",
        "reqId": "", "funcId": "COMM-005", "progId": "FileUploadServiceTest", "screenId": "",
        "purpose": "null 파일 업로드 시 예외 발생 검증",
        "procedure": "1. null로 upload 호출",
        "detail": "IllegalArgumentException 발생 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-COMM-006", "type": "기능", "section": "서비스구간",
        "major": "공통", "minor": "파일 업로드 서비스",
        "reqId": "", "funcId": "COMM-006", "progId": "FileUploadServiceTest", "screenId": "",
        "purpose": "서브 폴더 미지정 시 루트에 저장 검증",
        "procedure": "1. subFolder=null로 upload 호출",
        "detail": "결과 not null 검증 (루트 디렉토리 저장)",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 메일(MAIL) =====
    {
        "id": "UST-F-MAIL-001", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 서비스",
        "reqId": "", "funcId": "MAIL-001", "progId": "MailServiceTest", "screenId": "",
        "purpose": "메일 발송 성공 검증",
        "procedure": "1. MailRequest 생성\n2. send 호출",
        "detail": "mailSender.send(mimeMessage) verify",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-002", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 서비스",
        "reqId": "", "funcId": "MAIL-002", "progId": "MailServiceTest", "screenId": "",
        "purpose": "발신자 null 시 예외 발생 검증",
        "procedure": "1. sender=null인 MailRequest로 send 호출",
        "detail": "Exception 발생 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-003", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 컨트롤러",
        "reqId": "", "funcId": "MAIL-003", "progId": "MailControllerTest", "screenId": "",
        "purpose": "메일 발송 API 성공 검증",
        "procedure": "1. POST /api/mail/send 요청",
        "detail": "status 200, result='ok' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-004", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 컨트롤러",
        "reqId": "", "funcId": "MAIL-004", "progId": "MailControllerTest", "screenId": "",
        "purpose": "메일 발송 API 실패 검증",
        "procedure": "1. 메일 발송 예외 설정\n2. POST /api/mail/send 요청",
        "detail": "status 400, result='fail' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-005", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 알림 서비스",
        "reqId": "", "funcId": "MAIL-005", "progId": "MailNotificationServiceTest", "screenId": "",
        "purpose": "기본 수신자에게 알림 메일 발송 검증",
        "procedure": "1. sendNotification(title, subject, content) 호출",
        "detail": "MailRequest의 sender/receiver/title/subject/content 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-006", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 알림 서비스",
        "reqId": "", "funcId": "MAIL-006", "progId": "MailNotificationServiceTest", "screenId": "",
        "purpose": "지정 수신자에게 알림 메일 발송 검증",
        "procedure": "1. sendNotification(title, subject, content, receiver) 호출",
        "detail": "receiver가 지정 주소인지 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-007", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 알림 서비스",
        "reqId": "", "funcId": "MAIL-007", "progId": "MailNotificationServiceTest", "screenId": "",
        "purpose": "메일 발송 실패 시 예외 미전파 검증",
        "procedure": "1. 메일 발송 예외 설정\n2. sendNotification 호출",
        "detail": "assertDoesNotThrow 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-008", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 알림 서비스",
        "reqId": "", "funcId": "MAIL-008", "progId": "MailNotificationServiceTest", "screenId": "",
        "purpose": "본문 null인 경우 메일 발송 시도 검증",
        "procedure": "1. content=null로 sendNotification 호출",
        "detail": "content null 상태로 send 호출 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MAIL-009", "type": "기능", "section": "서비스구간",
        "major": "메일", "minor": "메일 알림 서비스",
        "reqId": "", "funcId": "MAIL-009", "progId": "MailNotificationServiceTest", "screenId": "",
        "purpose": "제목 null인 경우 메일 발송 시도 검증",
        "procedure": "1. subject=null로 sendNotification 호출",
        "detail": "subject null 상태로 send 호출 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 회원(MEMB) =====
    {
        "id": "UST-F-MEMB-001", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-001", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "이메일 중복 확인 — 존재하는 이메일",
        "procedure": "1. existsByEmail 호출 (등록된 이메일)",
        "detail": "true 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-002", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-002", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "이메일 중복 확인 — 존재하지 않는 이메일",
        "procedure": "1. existsByEmail 호출 (미등록 이메일)",
        "detail": "false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-003", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-003", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "실명인증 성공 검증",
        "procedure": "1. verifyIdentity('홍길동', '1990-01-01') 호출",
        "detail": "true 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-004", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-004", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "실명인증 실패 — 이름 누락",
        "procedure": "1. verifyIdentity('', '1990-01-01') 호출",
        "detail": "false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-005", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-005", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "실명인증 실패 — 생년월일 누락",
        "procedure": "1. verifyIdentity('홍길동', null) 호출",
        "detail": "false 반환 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-006", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 서비스",
        "reqId": "", "funcId": "MEMB-006", "progId": "MemberServiceTest", "screenId": "",
        "purpose": "회원 등록 성공 검증",
        "procedure": "1. create(member) 호출",
        "detail": "memberName, email 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-007", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-007", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 상세 조회 API 성공 검증",
        "procedure": "1. GET /api/member/detail?email=test@example.com 요청",
        "detail": "status 200, memberName/email 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-008", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-008", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 상세 조회 API — 존재하지 않는 이메일 시 404",
        "procedure": "1. GET /api/member/detail?email=unknown 요청",
        "detail": "status 404 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-009", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-009", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 프로필 수정 API 성공 검증",
        "procedure": "1. PUT /api/member/update-profile 요청",
        "detail": "status 200, success=true 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-010", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-010", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 프로필 수정 API — 회원 없음",
        "procedure": "1. PUT /api/member/update-profile 미등록 이메일 요청",
        "detail": "success=false, 에러 메시지 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-011", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-011", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "비밀번호 변경 API 성공 검증",
        "procedure": "1. PUT /api/member/change-password 요청",
        "detail": "status 200, success=true 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-012", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-012", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "비밀번호 변경 API — 현재 비밀번호 불일치",
        "procedure": "1. PUT /api/member/change-password 잘못된 현재 비밀번호 요청",
        "detail": "success=false, 에러 메시지 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-013", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-013", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "이메일 중복 확인 API — 중복 없음",
        "procedure": "1. GET /api/member/check-email?email=new 요청",
        "detail": "exists=false 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-014", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-014", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "이메일 중복 확인 API — 중복 있음",
        "procedure": "1. GET /api/member/check-email?email=test 요청",
        "detail": "exists=true 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-015", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-015", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "실명인증 API 성공 검증",
        "procedure": "1. POST /api/member/verify-identity 요청",
        "detail": "verified=true 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-016", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-016", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 등록 API 성공 검증",
        "procedure": "1. POST /api/member/register 요청",
        "detail": "status 200, memberName 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-MEMB-017", "type": "기능", "section": "서비스구간",
        "major": "회원", "minor": "회원 컨트롤러",
        "reqId": "", "funcId": "MEMB-017", "progId": "MemberControllerTest", "screenId": "",
        "purpose": "회원 등록 API — 이메일 중복 시 실패",
        "procedure": "1. POST /api/member/register 중복 이메일 요청",
        "detail": "status 400 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    # ===== 골프(GOLF) =====
    {
        "id": "UST-F-GOLF-001", "type": "기능", "section": "단위",
        "major": "골프장", "minor": "골프장 VO",
        "reqId": "", "funcId": "GOLF-001", "progId": "GolfCourseTest", "screenId": "",
        "purpose": "GolfCourse 기본 생성 및 필드 설정 검증",
        "procedure": "1. GolfCourse 생성 후 setter/getter 호출",
        "detail": "name, latitude, longitude, address, region, type 일치 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-002", "type": "기능", "section": "단위",
        "major": "골프장", "minor": "골프장 VO",
        "reqId": "", "funcId": "GOLF-002", "progId": "GolfCourseTest", "screenId": "",
        "purpose": "GolfCourse toString 출력 검증",
        "procedure": "1. GolfCourse 생성 후 toString 호출",
        "detail": "결과 not null 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-003", "type": "기능", "section": "단위",
        "major": "골프장", "minor": "골프장 VO",
        "reqId": "", "funcId": "GOLF-003", "progId": "GolfCourseTest", "screenId": "",
        "purpose": "GolfCourse equals 및 hashCode 검증",
        "procedure": "1. 동일 필드값의 GolfCourse 2개 생성\n2. equals/hashCode 비교",
        "detail": "equals true, hashCode 동일 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-004", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-004", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "JSON 파일 미존재 시 빈 리스트 초기화 검증",
        "procedure": "1. loadGolfCourses 호출 (파일 없음)\n2. retrieveAll 호출",
        "detail": "결과 not null, isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-005", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-005", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "전체 골프장 목록 조회 검증",
        "procedure": "1. 테스트 데이터 3건 주입\n2. retrieveAll 호출",
        "detail": "결과 크기 3 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-006", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-006", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "키워드 null 입력 시 전체 목록 반환 검증",
        "procedure": "1. searchByKeyword(null) 호출",
        "detail": "결과 크기 3 (전체) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-007", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-007", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "키워드 빈 문자열 입력 시 전체 목록 반환 검증",
        "procedure": "1. searchByKeyword('') 호출",
        "detail": "결과 크기 3 (전체) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-008", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-008", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "키워드 공백 문자열 입력 시 전체 목록 반환 검증",
        "procedure": "1. searchByKeyword('   ') 호출",
        "detail": "결과 크기 3 (전체) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-009", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-009", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "골프장 이름으로 키워드 검색 검증",
        "procedure": "1. searchByKeyword('남서울') 호출",
        "detail": "결과 1건, name='남서울CC' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-010", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-010", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "지역으로 키워드 검색 검증",
        "procedure": "1. searchByKeyword('제주') 호출",
        "detail": "결과 1건, name='제주오라CC' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-011", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-011", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "대소문자 무시 검색 검증",
        "procedure": "1. searchByKeyword('cc') 호출",
        "detail": "결과 3건 (모든 골프장 CC 포함) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-012", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-012", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "매칭 결과 없는 키워드 검색 검증",
        "procedure": "1. searchByKeyword('존재하지않는골프장') 호출",
        "detail": "결과 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-013", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 서비스",
        "reqId": "", "funcId": "GOLF-013", "progId": "GolfCourseServiceTest", "screenId": "",
        "purpose": "경기 지역 키워드로 복수 결과 검색 검증",
        "procedure": "1. searchByKeyword('경') 호출",
        "detail": "결과 2건 (경기, 경북) 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-014", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 컨트롤러",
        "reqId": "", "funcId": "GOLF-014", "progId": "GolfCourseControllerTest", "screenId": "",
        "purpose": "전체 골프장 목록 조회 — 인증된 사용자",
        "procedure": "1. 인증 세션 설정\n2. GET /api/golf-courses 요청",
        "detail": "status 200, 배열 크기 2, 골프장명 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-015", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 컨트롤러",
        "reqId": "", "funcId": "GOLF-015", "progId": "GolfCourseControllerTest", "screenId": "",
        "purpose": "전체 골프장 목록 조회 — 미인증 시 401 반환",
        "procedure": "1. 세션 없이 GET /api/golf-courses 요청",
        "detail": "status 401 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-016", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 컨트롤러",
        "reqId": "", "funcId": "GOLF-016", "progId": "GolfCourseControllerTest", "screenId": "",
        "purpose": "키워드 검색 — 인증된 사용자, 결과 있음",
        "procedure": "1. 인증 세션 설정\n2. GET /api/golf-courses/search?keyword=남서울 요청",
        "detail": "status 200, 배열 크기 1, name='남서울컨트리클럽' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-017", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 컨트롤러",
        "reqId": "", "funcId": "GOLF-017", "progId": "GolfCourseControllerTest", "screenId": "",
        "purpose": "키워드 검색 — 미인증 시 401 반환",
        "procedure": "1. 세션 없이 GET /api/golf-courses/search 요청",
        "detail": "status 401 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-018", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "골프장 컨트롤러",
        "reqId": "", "funcId": "GOLF-018", "progId": "GolfCourseControllerTest", "screenId": "",
        "purpose": "키워드 미입력 시 전체 목록 반환 검증",
        "procedure": "1. 인증 세션 설정\n2. GET /api/golf-courses/search (키워드 없음) 요청",
        "detail": "status 200, 배열 크기 2 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-019", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 컨트롤러",
        "reqId": "", "funcId": "GOLF-019", "progId": "LocationControllerTest", "screenId": "",
        "purpose": "역지오코딩 — 인증된 사용자, 정상 주소 반환",
        "procedure": "1. 인증 세션 설정\n2. GET /api/location/reverse-geocode 요청",
        "detail": "status 200, address 값 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-020", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 컨트롤러",
        "reqId": "", "funcId": "GOLF-020", "progId": "LocationControllerTest", "screenId": "",
        "purpose": "역지오코딩 — 미인증 시 401 반환",
        "procedure": "1. 세션 없이 GET /api/location/reverse-geocode 요청",
        "detail": "status 401 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-021", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 컨트롤러",
        "reqId": "", "funcId": "GOLF-021", "progId": "LocationControllerTest", "screenId": "",
        "purpose": "역지오코딩 — 주소 변환 실패 시 빈 문자열 반환",
        "procedure": "1. 인증 세션 설정\n2. 좌표 (0,0)으로 요청",
        "detail": "status 200, address='' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-022", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 서비스",
        "reqId": "", "funcId": "GOLF-022", "progId": "LocationServiceTest", "screenId": "",
        "purpose": "정상 응답 시 display_name 주소 반환 검증",
        "procedure": "1. MockRestServiceServer로 정상 JSON 응답 설정\n2. reverseGeocode 호출",
        "detail": "반환값 '서울특별시 강남구 역삼동, 대한민국' 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-023", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 서비스",
        "reqId": "", "funcId": "GOLF-023", "progId": "LocationServiceTest", "screenId": "",
        "purpose": "응답에 display_name 없을 때 빈 문자열 반환 검증",
        "procedure": "1. display_name 없는 JSON 응답 설정\n2. reverseGeocode 호출",
        "detail": "반환값 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-024", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 서비스",
        "reqId": "", "funcId": "GOLF-024", "progId": "LocationServiceTest", "screenId": "",
        "purpose": "외부 API 서버 오류 시 빈 문자열 반환 검증",
        "procedure": "1. 500 응답 설정\n2. reverseGeocode 호출",
        "detail": "반환값 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-025", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 서비스",
        "reqId": "", "funcId": "GOLF-025", "progId": "LocationServiceTest", "screenId": "",
        "purpose": "외부 API 404 응답 시 빈 문자열 반환 검증",
        "procedure": "1. 404 응답 설정\n2. reverseGeocode 호출",
        "detail": "반환값 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
    {
        "id": "UST-F-GOLF-026", "type": "기능", "section": "서비스구간",
        "major": "골프장", "minor": "위치 서비스",
        "reqId": "", "funcId": "GOLF-026", "progId": "LocationServiceTest", "screenId": "",
        "purpose": "빈 JSON 응답 시 빈 문자열 반환 검증",
        "procedure": "1. 빈 JSON '{}' 응답 설정\n2. reverseGeocode 호출",
        "detail": "반환값 isEmpty 검증",
        "date": "2026-04-20", "tester": "sinhuiyo", "result": "Pass", "reviewer": ""
    },
]

# 기존 샘플 데이터 행(3행) 삭제 — 새 데이터로 대체
ws.delete_rows(3, ws.max_row - 2)

# 데이터 행 작성
for idx, tc in enumerate(test_cases):
    row = idx + 3  # 3행부터 시작

    # 컬럼 매핑: B=2, C=3, ..., Q=17
    col_map = {
        2: tc["id"],        # B: 시험ID
        3: tc["type"],      # C: 시험유형
        4: tc["section"],   # D: 시험구분
        5: tc["major"],     # E: 시험항목(대분류)
        6: tc["minor"],     # F: 시험항목(중분류)
        7: tc["reqId"],     # G: 요구사항ID
        8: tc["funcId"],    # H: 기능ID
        9: tc["progId"],    # I: 프로그램ID
        10: tc["screenId"], # J: 화면ID
        11: tc["purpose"],  # K: 시험목적
        12: tc["procedure"],# L: 시험절차
        13: tc["detail"],   # M: 시험내역
        14: tc["date"],     # N: 시험일자
        15: tc["tester"],   # O: 시험자
        16: tc["result"],   # P: 시험결과
        17: tc["reviewer"], # Q: 시험결과 확인자
    }

    for col, value in col_map.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = copy.copy(data_font)
        cell.border = copy.copy(thin_border)
        # 시험절차(L), 시험내역(M), 시험목적(K)은 좌측 정렬, 나머지는 가운데 정렬
        if col in (11, 12, 13):
            cell.alignment = copy.copy(left_align)
        else:
            cell.alignment = copy.copy(center_align)

print(f"총 {len(test_cases)}건의 테스트 케이스를 작성했습니다.")

# 저장
wb.save("doc/단위테스트결과서.xlsx")
print("doc/단위테스트결과서.xlsx 저장 완료")
